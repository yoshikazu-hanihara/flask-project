# blueprints/export.py
# ─────────────────────────────────────────────
# ひな形 .xlsx に計算値を差し込んでダウンロードする
# 「結合セルだと書き込み不可」問題を set_value() で自動回避
# ─────────────────────────────────────────────

from flask import Blueprint, session, send_file, redirect, url_for, flash
from flask import current_app as app
from io import BytesIO
import datetime, os, openpyxl, json
from openpyxl.cell.cell import MergedCell
from db import get_connection

export_bp = Blueprint("export", __name__, url_prefix="/export")

# === 1. ひな形パス（テンプレは static/template/ に置く） ==========
TPL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),  # flask-project/
    "static", "template", "estimate_template.xlsx"
)

# === 2. Pythonキー ⇔ セル番地マッピング ==========================
CELL_MAP = {
    "sales_price":             "C12",
    "order_quantity":          "C13",
    "product_weight":          "C14",
    "mold_unit_price":         "C15",
    "mold_count":              "C16",
    "raw_material_cost_total": "F20",
    "manufacturing_cost_total":"F21",
    "sales_admin_cost_total":  "F22",
    "profit_amount":           "F24",
    "profit_amount_total":     "F25",
    "profit_ratio":            "F26",
}

# === 保存ユーティリティ =============================================
def _save_history(user_id: int, filename: str, data: dict):
    """Save exported excel info for the user."""
    conn = get_connection()
    with conn.cursor() as cursor:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS excel_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                filename VARCHAR(255) NOT NULL,
                data_json TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        cursor.execute(
            "INSERT INTO excel_history (user_id, filename, data_json) VALUES (%s, %s, %s)",
            (user_id, filename, json.dumps(data)),
        )
    conn.commit()
    conn.close()

# === 3. 結合セルでも安全に書き込むユーティリティ ================
def set_value(ws, coord: str, value):
    """
    coord が結合セルの途中でも左上セルへ代入。
    """
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        # 含まれる結合範囲を探す
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(rng.min_row, rng.min_col, value)
                break
    else:
        cell.value = value

# === 4. ワークブック生成 =========================================
def _build_workbook(data: dict) -> BytesIO:
    wb = openpyxl.load_workbook(TPL_PATH)
    ws = wb.active               # 見積書シートは 1 枚目想定

    # 指定セルへ値を流し込む
    for key, cell in CELL_MAP.items():
        if key in data:
            set_value(ws, cell, data[key])

    # 発行日（任意）
    set_value(ws, "H3", datetime.date.today().strftime("%Y/%m/%d"))

    # ── 式再計算フラグ─────────────
    if hasattr(wb, "calculation") and wb.calculation is not None:
        # openpyxl ≥ 3.1
        wb.calculation.fullCalcOnLoad = True
    elif hasattr(wb, "calc_properties") and wb.calc_properties is not None:
        # openpyxl 3.0 系
        wb.calc_properties.fullCalcOnLoad = True
    # ────────────────────────────────────────

    # メモリへ保存
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# === 5. ファイル名ユーティリティ ================================
def _make_filename() -> str:
    return f"見積書_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx"

# === 6. ルート：ダウンロード ===================================
@export_bp.route("/excel")
def download_excel():
    data = session.get("dashboard_data")
    if not data:
        flash("先に見積りを計算してください。")
        return redirect(url_for("dashboard.dashboard"))

    bio = _build_workbook(data)
    filename = _make_filename()

    # -- サーバー側へユーザー単位で保存 ---------------------
    export_dir = os.path.join(app.root_path, "exports")
    os.makedirs(export_dir, exist_ok=True)
    if "user_id" in session:
        user_dir = os.path.join(export_dir, str(session["user_id"]))
        os.makedirs(user_dir, exist_ok=True)
        with open(os.path.join(user_dir, filename), "wb") as f:
            f.write(bio.getbuffer())
        _save_history(session["user_id"], filename, data)
    else:
        with open(os.path.join(export_dir, filename), "wb") as f:
            f.write(bio.getbuffer())

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
